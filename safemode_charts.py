#!/usr/bin/env python3
"""Generate SVG donut charts from SafeMode analysis XLSX."""

import argparse
import base64
import math
import sys
from pathlib import Path

from openpyxl import load_workbook

from everpure_colors import (
    ASH_GRAY, BASIL_GREEN, CINNAMON_BROWN, MINT_GREEN, MOSS_GREEN,
    PURE_ORANGE, QUARTZ_PINK, STONE_GRAY, WALNUT_BROWN,
)

# ---------------------------------------------------------------------------
# Layout constants
# ---------------------------------------------------------------------------

OUTER_R     = 90
INNER_R     = 50
MIN_ARC_RAD = math.radians(3)            # minimum visible slice angle

CARD_RX   = 12
ACCENT_H  = 3
PAD_X     = 40     # horizontal padding (each side)
PAD_TOP   = 28     # from accent-bar bottom to first text
SINGLE_W  = 360    # card width — single mode
DUAL_W    = 700    # card width — both mode
COL_W     = 280    # chart column width
COL_GAP   = 60     # gap between dual columns
ROW_H     = 22     # breakdown row height
BOT_PAD   = 28     # padding below breakdown

FONT = "PureSans, Arial, sans-serif"

# ---------------------------------------------------------------------------
# Risk level ordering / metadata
# ---------------------------------------------------------------------------

RISK_ORDER = ["PROTECTED", "PROTECTED (NOT BP)", "AT RISK", "CRITICAL", "EXCLUDED"]

SLICE_COLORS = {
    "PROTECTED":          BASIL_GREEN,     # #5A6359
    "PROTECTED (NOT BP)": MOSS_GREEN,      # #8FA596
    "AT RISK":            CINNAMON_BROWN,  # #BD673D
    "CRITICAL":           QUARTZ_PINK,     # #DEA193
    "EXCLUDED":           STONE_GRAY,      # #D0C8BA
}

RISK_LABEL = {
    "PROTECTED":         "Protected",
    "PROTECTED (NOT BP)":"Protected (Not BP)",
    "AT RISK":           "At Risk",
    "CRITICAL":          "Critical",
    "EXCLUDED":          "Excluded",
}

# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------

def _derive_risk(excl, safemode, snap_window: int) -> str:
    if str(excl or "").strip().lower() == "x":
        return "EXCLUDED"
    if safemode == "Yes":
        return "PROTECTED" if snap_window >= 14 else "PROTECTED (NOT BP)"
    return "AT RISK" if snap_window > 0 else "CRITICAL"


def _parse_cap(s) -> float:
    """Parse '9.51T', '225.88M', '1.93T', etc. to TB."""
    s = str(s or "").strip()
    if not s or s in ("-", "0.00", "0"):
        return 0.0
    import re
    m = re.match(r"^([\d.]+)\s*([TGMKtgmk]?)$", s)
    if not m:
        return 0.0
    v = float(m.group(1))
    u = m.group(2).upper()
    return {"T": v, "G": v / 1024, "M": v / 1024 ** 2, "K": v / 1024 ** 3, "": v}[u]


def _fmt_cap(tb: float) -> str:
    if tb >= 1:
        return f"{tb:.2f} TB"
    if tb >= 1 / 1024:
        return f"{tb * 1024:.2f} GB"
    return f"{tb * 1024 ** 2:.2f} MB"


def _fmt_pct(pct: float) -> str:
    if pct == 0:
        return "0%"
    if pct < 1:
        return "&lt;1%"
    return f"{round(pct)}%"


def load_data(xlsx_path: Path, only_arrays: list[str] | None) -> dict[str, dict]:
    """Return {array: {risk_level: {count, capacity}}} ordered by array appearance."""
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]

    def ci(name):
        return headers.index(name)

    ai, ei, smi, swi, effi = (
        ci("Array"), ci("Exclude"), ci("In SafeMode PGroup"),
        ci("Max Snap Window"), ci("Effective Used"),
    )

    arrays: dict[str, dict] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        arr = row[ai]
        if not arr:
            continue
        if only_arrays and arr not in only_arrays:
            continue

        risk = _derive_risk(row[ei], row[smi], int(row[swi] or 0))
        cap = _parse_cap(row[effi])

        if arr not in arrays:
            arrays[arr] = {r: {"count": 0, "cap": 0.0} for r in RISK_ORDER}
        arrays[arr][risk]["count"] += 1
        arrays[arr][risk]["cap"] += cap

    return arrays


# ---------------------------------------------------------------------------
# Font embedding
# ---------------------------------------------------------------------------

def _load_font_css(font_dir: Path) -> str:
    weights = [
        ("Regular", 400),
        ("SemiBold", 600),
        ("Bold", 700),
    ]
    css = ""
    for name, weight in weights:
        p = font_dir / f"PureSans-{name}.ttf"
        if p.exists():
            b64 = base64.b64encode(p.read_bytes()).decode()
            css += (
                f"@font-face{{font-family:'PureSans';font-weight:{weight};"
                f"src:url('data:font/truetype;base64,{b64}');}}\n"
            )
    return css


# ---------------------------------------------------------------------------
# SVG helpers
# ---------------------------------------------------------------------------

def _t(x, y, text, size, fill, weight=400, anchor="start", spacing="normal"):
    sp = f' letter-spacing="{spacing}"' if spacing != "normal" else ""
    return (
        f'<text x="{x}" y="{y}" font-family="{FONT}" font-size="{size}" '
        f'font-weight="{weight}" fill="{fill}" text-anchor="{anchor}"'
        f' dominant-baseline="auto"{sp}>{text}</text>'
    )


def _line(x1, y1, x2, y2, stroke, w=1):
    return f'<line x1="{x1}" y1="{y1}" x2="{x2}" y2="{y2}" stroke="{stroke}" stroke-width="{w}"/>'


# ---------------------------------------------------------------------------
# Donut chart
# ---------------------------------------------------------------------------

def _donut_slices(cx, cy, draw_buckets: list[tuple]) -> list[str]:
    """
    Filled pie slices from center + white inner circle for the donut hole.
    draw_buckets: sorted largest→smallest so smaller slices render on top.
    Straight radial edges between slices (no rounded ends).
    """
    total = sum(v for _, v in draw_buckets)
    if total == 0:
        return [
            f'<circle cx="{cx}" cy="{cy}" r="{OUTER_R}" fill="{STONE_GRAY}"/>',
            f'<circle cx="{cx}" cy="{cy}" r="{INNER_R}" fill="#FFFFFF"/>',
        ]

    parts = []
    theta = -math.pi / 2  # start from 12 o'clock

    for risk, value in draw_buckets:
        sweep = (value / total) * 2 * math.pi
        color = SLICE_COLORS[risk]

        if sweep >= 2 * math.pi - 0.001:
            parts.append(f'<circle cx="{cx}" cy="{cy}" r="{OUTER_R}" fill="{color}"/>')
        elif sweep >= MIN_ARC_RAD:
            x1 = cx + OUTER_R * math.cos(theta)
            y1 = cy + OUTER_R * math.sin(theta)
            x2 = cx + OUTER_R * math.cos(theta + sweep)
            y2 = cy + OUTER_R * math.sin(theta + sweep)
            large = 1 if sweep > math.pi else 0
            parts.append(
                f'<path d="M {cx:.1f} {cy:.1f} L {x1:.3f} {y1:.3f} '
                f'A {OUTER_R} {OUTER_R} 0 {large} 1 {x2:.3f} {y2:.3f} Z" '
                f'fill="{color}" stroke="#FFFFFF" stroke-width="1.5"/>'
            )
        theta += sweep

    # White center hole
    parts.append(f'<circle cx="{cx}" cy="{cy}" r="{INNER_R}" fill="#FFFFFF"/>')
    return parts


def _center_text(cx, cy, protected_count, total, field) -> list[str]:
    """Render center percentage + 'protected' label."""
    if total == 0:
        return [_t(cx, cy + 8, "N/A", 20, STONE_GRAY, weight=700, anchor="middle")]
    pct = protected_count / total * 100
    pct_str = _fmt_pct(pct)
    color = MINT_GREEN if pct >= 80 else PURE_ORANGE if pct >= 50 else QUARTZ_PINK
    return [
        _t(cx, cy + 8, pct_str, 28, color, weight=700, anchor="middle"),
        _t(cx, cy + 22, "protected", 10, WALNUT_BROWN, anchor="middle"),
    ]


# ---------------------------------------------------------------------------
# Breakdown table
# ---------------------------------------------------------------------------

def _breakdown(col_x, col_y, buckets: list[tuple], total, fmt) -> list[str]:
    """
    Render breakdown rows for visible (non-zero) buckets + total row.
    col_x/y = top-left of the breakdown area.
    """
    parts = []
    y = col_y

    for risk, value in buckets:
        pct = value / total * 100 if total > 0 else 0
        color = SLICE_COLORS[risk]
        label = RISK_LABEL[risk]
        val_str = f"{fmt(value)} ({_fmt_pct(pct)})"

        mid_y = y + ROW_H // 2
        base_y = mid_y + 5

        parts.append(f'<circle cx="{col_x + 4}" cy="{mid_y}" r="4" fill="{color}"/>')
        parts.append(_t(col_x + 14, base_y, label, 12, ASH_GRAY))
        parts.append(_t(col_x + COL_W, base_y, val_str, 12, ASH_GRAY, weight=700, anchor="end"))
        y += ROW_H
        parts.append(_line(col_x, y, col_x + COL_W, y, STONE_GRAY))

    # Total row — separator + bold total
    sep_y = y + 4
    parts.append(_line(col_x, sep_y, col_x + COL_W, sep_y, ASH_GRAY))
    base_y = y + ROW_H // 2 + 5
    parts.append(_t(col_x, base_y, "Total", 12, ASH_GRAY, weight=700))
    parts.append(_t(col_x + COL_W, base_y, fmt(total), 12, ASH_GRAY, weight=700, anchor="end"))

    return parts


# ---------------------------------------------------------------------------
# Card rendering
# ---------------------------------------------------------------------------

def _render_chart_col(
    col_x: int, card_y: int, y_pos: dict,
    display_buckets: list[tuple], draw_buckets: list[tuple],
    total, protected_count, title: str, fmt, parts: list,
):
    """Render one chart column (title + donut + breakdown) into parts list."""
    cx = col_x + COL_W // 2
    cy = card_y + y_pos["donut_cy"]

    # Chart title
    parts.append(_t(
        cx, card_y + y_pos["chart_title_base"],
        title, 14, WALNUT_BROWN, weight=700, anchor="middle", spacing="0.5",
    ))

    # Donut (draw order: largest slice first)
    parts += _donut_slices(cx, cy, draw_buckets)
    parts += _center_text(cx, cy, protected_count, total, fmt)

    # Breakdown (display order: RISK_ORDER — best to worst)
    bkdn_y = card_y + y_pos["bkdn_top"]
    parts += _breakdown(col_x, bkdn_y, display_buckets, total, fmt)


def _build_y(n_bkdn_rows: int, has_excluded: bool = False) -> dict:
    """Pre-compute Y positions (relative to card top) given breakdown row count."""
    y = {}
    y["name_base"]        = ACCENT_H + PAD_TOP + 22   # 53  — array name baseline
    y["divider"]          = y["name_base"] + 12        # 65  — header divider
    cur                   = y["divider"] + 2 + 18      # 85  — chart section start
    y["chart_title_base"] = cur + 14                   # 99
    y["donut_top"]        = y["chart_title_base"] + 14 # 113
    y["donut_cy"]         = y["donut_top"] + OUTER_R   # 203
    y["bkdn_top"]         = y["donut_top"] + 2 * OUTER_R + 24  # 317
    bkdn_end              = y["bkdn_top"] + (n_bkdn_rows + 1) * ROW_H
    if has_excluded:
        y["footnote_top"] = bkdn_end + 8
        y["card_h"]       = y["footnote_top"] + ROW_H + BOT_PAD
    else:
        y["card_h"]       = bkdn_end + BOT_PAD
    return y


def render_card(
    array_name: str,
    data: dict,
    mode: str,
    font_css: str,
) -> str:
    """Generate a complete SVG card for one array."""

    # Excluded volumes are always removed from chart/breakdown; shown only as footnote
    def _display(field: str) -> list[tuple]:
        return [
            (r, data[r][field])
            for r in RISK_ORDER
            if data[r][field] > 0 and r != "EXCLUDED"
        ]

    def _draw(field: str) -> list[tuple]:
        return _display(field)  # RISK_ORDER clockwise, matching legend

    count_display = _display("count")
    count_draw    = _draw("count")
    cap_display   = _display("cap")
    cap_draw      = _draw("cap")

    count_total = sum(v for _, v in count_display)
    cap_total   = sum(v for _, v in cap_display)

    # Center pct = Protected + Protected (Not BP) combined
    protected_count = data["PROTECTED"]["count"] + data["PROTECTED (NOT BP)"]["count"]
    protected_cap   = data["PROTECTED"]["cap"]   + data["PROTECTED (NOT BP)"]["cap"]

    excl_count = data["EXCLUDED"]["count"]
    excl_cap   = data["EXCLUDED"]["cap"]

    n_bkdn = len(count_display)
    y = _build_y(n_bkdn, has_excluded=(excl_count > 0))

    card_w = DUAL_W if mode == "both" else SINGLE_W
    card_h = y["card_h"]

    # Header stats — non-excluded totals only; footnote handles exclusion callout
    stats = f"{count_total} volumes  |  {_fmt_cap(cap_total)} effective"

    # SVG dimensions — card fills the SVG with a small margin for shadow
    M = 8  # margin for shadow bleed
    svg_w = card_w + 2 * M
    svg_h = card_h + 2 * M
    cx0 = M   # card origin X within SVG
    cy0 = M   # card origin Y within SVG

    parts: list[str] = []

    # --- SVG header ---
    parts.append(
        f'<svg xmlns="http://www.w3.org/2000/svg" '
        f'width="{svg_w}" height="{svg_h}" '
        f'viewBox="0 0 {svg_w} {svg_h}" '
        f'shape-rendering="geometricPrecision">'
    )

    # --- Defs ---
    defs = []
    if font_css:
        defs.append(f"<style>{font_css}</style>")
    defs.append(
        '<filter id="s" x="-5%" y="-5%" width="110%" height="115%">'
        '<feDropShadow dx="0" dy="2" stdDeviation="4" '
        'flood-color="#000000" flood-opacity="0.08"/>'
        "</filter>"
    )
    # ClipPath for accent bar (matches card rounded corners)
    defs.append(
        f'<clipPath id="c">'
        f'<rect x="{cx0}" y="{cy0}" width="{card_w}" height="{card_h}" rx="{CARD_RX}"/>'
        f"</clipPath>"
    )
    parts.append("<defs>" + "".join(defs) + "</defs>")

    # --- Card background (with shadow) ---
    parts.append(
        f'<rect x="{cx0}" y="{cy0}" width="{card_w}" height="{card_h}" '
        f'rx="{CARD_RX}" fill="#FFFFFF" filter="url(#s)"/>'
    )

    # --- Accent bar (clipped to card corners) ---
    parts.append(
        f'<rect x="{cx0}" y="{cy0}" width="{card_w}" height="{CARD_RX + ACCENT_H}" '
        f'rx="{CARD_RX}" fill="{PURE_ORANGE}" clip-path="url(#c)"/>'
    )
    # Cover the rounded bottom of the accent bar rect with white
    parts.append(
        f'<rect x="{cx0}" y="{cy0 + ACCENT_H}" width="{card_w}" height="{CARD_RX}" '
        f'fill="#FFFFFF" clip-path="url(#c)"/>'
    )

    # --- Header ---
    right_x = cx0 + card_w - PAD_X
    parts.append(_t(cx0 + PAD_X, cy0 + y["name_base"], array_name, 20, ASH_GRAY, weight=700))
    parts.append(_t(right_x, cy0 + y["name_base"], stats, 12, WALNUT_BROWN, anchor="end"))

    # Divider line spanning card inner width
    div_y = cy0 + y["divider"]
    parts.append(_line(cx0 + PAD_X, div_y, right_x, div_y, ASH_GRAY))

    # --- Chart column(s) ---
    if mode == "both":
        left_col_x  = cx0 + PAD_X
        right_col_x = cx0 + PAD_X + COL_W + COL_GAP

        _render_chart_col(
            left_col_x, cy0, y, count_display, count_draw,
            count_total, protected_count, "BY VOLUME COUNT",
            lambda v: f"{int(v)} vols", parts,
        )
        _render_chart_col(
            right_col_x, cy0, y, cap_display, cap_draw,
            cap_total, protected_cap, "BY EFFECTIVE CAPACITY",
            _fmt_cap, parts,
        )
    else:
        col_x = cx0 + PAD_X
        if mode == "count":
            _render_chart_col(
                col_x, cy0, y, count_display, count_draw,
                count_total, protected_count, "BY VOLUME COUNT",
                lambda v: f"{int(v)} vols", parts,
            )
        else:
            _render_chart_col(
                col_x, cy0, y, cap_display, cap_draw,
                cap_total, protected_cap, "BY EFFECTIVE CAPACITY",
                _fmt_cap, parts,
            )

    # --- Exclusion footnote ---
    if excl_count > 0:
        fn_y = cy0 + y["footnote_top"] + ROW_H // 2 + 4
        fn_text = f"ℹ {excl_count} volumes excluded ({_fmt_cap(excl_cap)})"
        parts.append(_t(cx0 + PAD_X, fn_y, fn_text, 11, WALNUT_BROWN))

    parts.append("</svg>")
    return "\n".join(parts)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main() -> None:
    ap = argparse.ArgumentParser(description="Generate SVG protection charts from SafeMode XLSX")
    ap.add_argument("--input", required=True, help="SafeMode report .xlsx")
    ap.add_argument("--output-dir", required=True, help="Directory for output SVG files")
    ap.add_argument("--mode", choices=["count", "capacity", "both"], default="both")
    ap.add_argument("--arrays", help="Comma-separated list of array names to include")
    ap.add_argument("--min-pct", type=float, default=0,
                    help="Merge slices below this %% into Other (default: 0)")
    ap.add_argument("--font-dir", help="Directory containing PureSans .ttf files")
    args = ap.parse_args()

    xlsx_path = Path(args.input)
    if not xlsx_path.exists():
        print(f"ERROR: input file not found: {xlsx_path}", file=sys.stderr)
        sys.exit(1)

    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    only_arrays = [a.strip() for a in args.arrays.split(",")] if args.arrays else None

    # Load font
    font_dir = Path(args.font_dir) if args.font_dir else Path(__file__).parent / "fonts" / "Pure_Sans"
    font_css = _load_font_css(font_dir) if font_dir.exists() else ""
    if not font_css:
        import logging
        logging.warning("PureSans font not found at %s — falling back to Arial", font_dir)

    # Load data
    arrays = load_data(xlsx_path, only_arrays)
    if not arrays:
        print("No data found in input file.", file=sys.stderr)
        sys.exit(1)

    # Generate one SVG per array
    for array_name, data in arrays.items():
        svg = render_card(array_name, data, args.mode, font_css)
        safe_name = array_name.replace("/", "_").replace(":", "_")
        out_path = output_dir / f"{safe_name}_{args.mode}.svg"
        out_path.write_text(svg)
        print(f"  {out_path}")

    print(f"Generated {len(arrays)} SVG(s) → {output_dir}/")


if __name__ == "__main__":
    main()
