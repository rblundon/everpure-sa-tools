#!/usr/bin/env python3
"""Pure Storage FlashArray SafeMode Parser"""

import argparse
import logging
import re
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
log = logging.getLogger(__name__)

_DASH_RE = re.compile(r"^-{10,}")
_TIMESTAMP_RE = re.compile(
    r"^(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s+\d+\s+\d+:\d+:\d+"
)


def _skip_preamble(lines: list[str]) -> list[str]:
    for i, line in enumerate(lines):
        s = line.strip()
        if s and not _DASH_RE.match(s) and not _TIMESTAMP_RE.match(s):
            return lines[i:]
    return []


def _col_positions(header: str) -> dict[str, int]:
    """Return {column_name: start_pos} by splitting on runs of 2+ spaces."""
    result: dict[str, int] = {}
    pos = 0
    for part in re.split(r"(\s{2,})", header):
        if re.match(r"\s{2,}", part):
            pos += len(part)
        else:
            name = part.strip()
            if name:
                result[name] = pos
            pos += len(part)
    return result


def _extract(line: str, positions: dict[str, int]) -> dict[str, str]:
    """Extract values from a fixed-width line given column start positions."""
    ordered = sorted(positions.items(), key=lambda x: x[1])
    result: dict[str, str] = {}
    for idx, (col, start) in enumerate(ordered):
        end = ordered[idx + 1][1] if idx + 1 < len(ordered) else len(line)
        result[col] = line[start:end].strip() if len(line) > start else ""
    return result


def _retention_days(all_for: str) -> int:
    """Convert '3d' or '2w' string to integer days."""
    m = re.match(r"^(\d+)([dw])$", all_for.strip(), re.IGNORECASE)
    if not m:
        if all_for.strip():
            log.warning("Unrecognized retention format '%s', defaulting to 0", all_for)
        return 0
    v = int(m.group(1))
    return v * 7 if m.group(2).lower() == "w" else v


# ---------------------------------------------------------------------------
# Per-file parsers
# ---------------------------------------------------------------------------

def parse_volumes(lines: list[str]) -> list[dict]:
    """purevol-list.out → [{name, provisioned, effective_used}]"""
    content = _skip_preamble(lines)
    if not content:
        return []
    pos = _col_positions(content[0])
    volumes = []
    for line in content[1:]:
        if not line.strip():
            continue
        cols = _extract(line, pos)
        name = cols.get("Name", "")
        if not name or name.startswith("(total)"):
            continue
        volumes.append(
            {
                "name": name,
                "provisioned": cols.get("Size", ""),
                "effective_used": cols.get("Total", ""),
            }
        )
    return volumes


def parse_pgroup_list(lines: list[str]) -> dict[str, dict]:
    """
    purepgroup-list.out → {pgroup: {host_groups: [...], volumes: [...]}}
    Handles multi-line volume continuation rows.
    """
    content = _skip_preamble(lines)
    if not content:
        return {}
    pos = _col_positions(content[0])
    name_end = min(v for k, v in pos.items() if k != "Name")

    pgroups: dict[str, dict] = {}
    current: str | None = None

    for line in content[1:]:
        if not line.strip():
            continue
        name_field = line[:name_end].strip() if len(line) > name_end else ""
        if name_field:
            cols = _extract(line, pos)
            current = cols["Name"]
            pgroups[current] = {"host_groups": [], "volumes": []}
            hg = cols.get("Host Groups", "")
            if hg and hg != "-":
                pgroups[current]["host_groups"].append(hg)
            vol = cols.get("Volumes", "")
            if vol and vol != "-":
                pgroups[current]["volumes"].append(vol)
        else:
            if current is None:
                continue
            vol_start = pos.get("Volumes", 0)
            vol = line[vol_start:].strip() if len(line) > vol_start else ""
            if vol and vol != "-":
                pgroups[current]["volumes"].append(vol)

    return pgroups


def parse_retention_lock(lines: list[str]) -> dict[str, str]:
    """purepgroup-list-retention-lock-pending.out → {pgroup: retention_lock}"""
    content = _skip_preamble(lines)
    if not content:
        return {}
    pos = _col_positions(content[0])
    result: dict[str, str] = {}
    for line in content[1:]:
        if not line.strip():
            continue
        cols = _extract(line, pos)
        name = cols.get("Name", "")
        if name:
            result[name] = cols.get("Retention Lock", "unlocked")
    return result


def parse_retention(lines: list[str]) -> dict[str, int]:
    """
    purepgroup-list-retention.out → {pgroup: total_source_days}
    Uses source rows only. All For + Days = total window.
    """
    content = _skip_preamble(lines)
    if not content:
        return {}
    pos = _col_positions(content[0])
    name_end = min(v for k, v in pos.items() if k != "Name")

    result: dict[str, int] = {}
    current: str | None = None
    current_done = False

    for line in content[1:]:
        if not line.strip():
            continue
        name_field = line[:name_end].strip() if len(line) > name_end else ""
        if name_field:
            current = name_field
            current_done = False
        if current is None:
            continue
        cols = _extract(line, pos)
        array_val = cols.get("Array", "").lower()
        if array_val == "source" and not current_done:
            all_for = cols.get("All For", "0d")
            try:
                days = int(cols.get("Days", "0") or 0)
            except ValueError:
                days = 0
            result[current] = _retention_days(all_for) + days
            current_done = True

    return result


def parse_hgroup_connect(lines: list[str]) -> dict[str, list[str]]:
    """
    purehgroup-list-connect.out (fixed-width or CSV) → {hgroup: [volumes]}
    """
    content = _skip_preamble(lines)
    if not content:
        return {}
    header = content[0]
    result: dict[str, list[str]] = {}

    if "," in header:
        # CSV: Name,Lun,Vol
        for line in content[1:]:
            if not line.strip():
                continue
            parts = line.split(",")
            if len(parts) >= 3:
                hg, vol = parts[0].strip(), parts[2].strip()
                if hg and vol:
                    result.setdefault(hg, []).append(vol)
    else:
        pos = _col_positions(header)
        for line in content[1:]:
            if not line.strip():
                continue
            cols = _extract(line, pos)
            hg = cols.get("Name", "")
            vol = cols.get("Vol", "")
            if hg and vol:
                result.setdefault(hg, []).append(vol)

    return result


# ---------------------------------------------------------------------------
# Array-level orchestration
# ---------------------------------------------------------------------------

def _read(folder: Path, filename: str) -> list[str] | None:
    path = folder / filename
    if not path.exists():
        log.warning("Missing file: %s", path)
        return None
    return path.read_text().splitlines()


def parse_array(folder: Path) -> list[dict]:
    """Parse one array folder and return a list of volume row dicts."""
    array_name = folder.name

    vol_lines = _read(folder, "purevol-list.out")
    if not vol_lines:
        log.warning("No volume data for %s, skipping", array_name)
        return []

    volumes = parse_volumes(vol_lines)
    if not volumes:
        log.warning("No volumes found in %s/purevol-list.out", array_name)
        return []

    pg_lines = _read(folder, "purepgroup-list.out")
    pgroups = parse_pgroup_list(pg_lines) if pg_lines else {}

    rl_lines = _read(folder, "purepgroup-list-retention-lock-pending.out")
    retention_lock = parse_retention_lock(rl_lines) if rl_lines else {}

    ret_lines = _read(folder, "purepgroup-list-retention.out")
    retention_days = parse_retention(ret_lines) if ret_lines else {}

    hg_lines = _read(folder, "purehgroup-list-connect.out")
    hgroup_vols = parse_hgroup_connect(hg_lines) if hg_lines else {}

    vol_names = {v["name"] for v in volumes}

    # Warn about unresolvable host group memberships
    if not hg_lines:
        for pg_name, pg_data in pgroups.items():
            if pg_data["host_groups"]:
                log.warning(
                    "Cannot resolve host group membership for pgroup '%s' "
                    "(host groups: %s): purehgroup-list-connect.out missing. "
                    "Those volumes will appear unprotected.",
                    pg_name,
                    pg_data["host_groups"],
                )

    # Resolve pgroup → volume set (direct + via host group)
    pg_volumes: dict[str, set[str]] = {}
    for pg_name, pg_data in pgroups.items():
        members: set[str] = set()
        for v in pg_data["volumes"]:
            if v in vol_names:
                members.add(v)
        for hg in pg_data["host_groups"]:
            hg_vol_list = hgroup_vols.get(hg, [])
            if not hg_vol_list and hg_lines:
                log.warning(
                    "Host group '%s' in pgroup '%s' not found in "
                    "purehgroup-list-connect.out",
                    hg,
                    pg_name,
                )
            for v in hg_vol_list:
                if v in vol_names:
                    members.add(v)
        pg_volumes[pg_name] = members

    # Build per-volume protection summary
    vol_protection: dict[str, dict] = {}
    for pg_name, members in pg_volumes.items():
        is_safemode = retention_lock.get(pg_name, "unlocked") == "ratcheted"
        snap_window = retention_days.get(pg_name, 0)
        for vol in members:
            vp = vol_protection.setdefault(vol, {"max_snap_window": 0, "in_safemode_pgroup": False})
            vp["max_snap_window"] = max(vp["max_snap_window"], snap_window)
            if is_safemode:
                vp["in_safemode_pgroup"] = True

    rows = []
    for vol in volumes:
        vname = vol["name"]
        prot = vol_protection.get(vname, {"max_snap_window": 0, "in_safemode_pgroup": False})
        rows.append(
            {
                "Array": array_name,
                "Volume Name": vname,
                "Provisioned": vol["provisioned"],
                "Effective Used": vol["effective_used"],
                "Max Snap Window": prot["max_snap_window"],
                "In SafeMode PGroup": "Yes" if prot["in_safemode_pgroup"] else "No",
                "Exclude": "",
                "Reason": "",
            }
        )
    return rows


# ---------------------------------------------------------------------------
# Persist logic
# ---------------------------------------------------------------------------

def load_existing(path: Path) -> dict[tuple, tuple]:
    """Load (Array, Volume Name) → (Exclude, Reason) from an existing workbook."""
    if not path or not path.exists():
        return {}
    wb = load_workbook(path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    try:
        ai = headers.index("Array")
        vi = headers.index("Volume Name")
        ei = headers.index("Exclude")
        ri = headers.index("Reason")
    except ValueError:
        log.warning("Existing spreadsheet missing expected columns; ignoring persist data")
        return {}
    result: dict[tuple, tuple] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        arr, vol = row[ai], row[vi]
        exc, rea = row[ei] or "", row[ri] or ""
        if arr and vol:
            result[(arr, vol)] = (exc, rea)
    return result


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------

_DARK_BG = "3C3C3B"
_GREEN = "C8E6C0"
_YELLOW_GREEN = "E8F5C8"
_YELLOW = "FFF3C4"
_RED = "F8C4C4"
_GRAY = "E0E0E0"

COLUMNS = [
    "Array", "Volume Name", "Provisioned", "Effective Used",
    "Max Snap Window", "In SafeMode PGroup", "Exclude", "Reason", "Risk Level",
]
# Column letters (1-indexed) for use in formulas
_COL = {name: get_column_letter(i + 1) for i, name in enumerate(COLUMNS)}


def write_xlsx(rows: list[dict], output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "SafeMode Analysis"

    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor=_DARK_BG)
    data_font = Font(name="Arial", size=10)

    for col_idx, col_name in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left")

    exc_col = _COL["Exclude"]
    sm_col = _COL["In SafeMode PGroup"]
    win_col = _COL["Max Snap Window"]

    for row_idx, row in enumerate(rows, 2):
        for col_idx, col_name in enumerate(COLUMNS, 1):
            if col_name == "Risk Level":
                value = (
                    f'=IF({exc_col}{row_idx}="x","EXCLUDED",'
                    f'IF({sm_col}{row_idx}="No",'
                    f'IF({win_col}{row_idx}=0,"CRITICAL","AT RISK"),'
                    f'IF({win_col}{row_idx}>=14,"PROTECTED","PROTECTED (NOT BP)")))'
                )
            else:
                value = row.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    risk_col = _COL["Risk Level"]
    last_row = len(rows) + 1
    risk_range = f"{risk_col}2:{risk_col}{last_row}"
    for value, color in [
        ("PROTECTED", _GREEN),
        ("PROTECTED (NOT BP)", _YELLOW_GREEN),
        ("AT RISK", _YELLOW),
        ("CRITICAL", _RED),
        ("EXCLUDED", _GRAY),
    ]:
        ws.conditional_formatting.add(
            risk_range,
            CellIsRule(
                operator="equal",
                formula=[f'"{value}"'],
                fill=PatternFill(fill_type="solid", fgColor=color),
            ),
        )

    for col_idx, col_name in enumerate(COLUMNS, 1):
        col_letter = get_column_letter(col_idx)
        max_width = len(col_name)
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_width = max(max_width, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_width + 2, 60)

    wb.save(output_path)
    log.info("Saved report to %s", output_path)


# ---------------------------------------------------------------------------
# CLI entry point
# ---------------------------------------------------------------------------

def main() -> None:
    ap = argparse.ArgumentParser(description="Pure Storage FlashArray SafeMode Parser")
    ap.add_argument("--input-dir", required=True, help="Root folder with one subfolder per array")
    ap.add_argument("--output", required=True, help="Output .xlsx path")
    ap.add_argument("--existing", help="Existing spreadsheet to persist Exclude/Reason from")
    args = ap.parse_args()

    input_dir = Path(args.input_dir)
    output_path = Path(args.output)

    if not input_dir.is_dir():
        log.error("Input directory not found: %s", input_dir)
        sys.exit(1)

    persist = load_existing(Path(args.existing)) if args.existing else {}

    array_folders = sorted(d for d in input_dir.iterdir() if d.is_dir())
    if not array_folders:
        log.error("No array subfolders found in %s", input_dir)
        sys.exit(1)

    all_rows: list[dict] = []
    for folder in array_folders:
        all_rows.extend(parse_array(folder))

    if not all_rows:
        log.error("No volume data found across all arrays")
        sys.exit(1)

    for row in all_rows:
        key = (row["Array"], row["Volume Name"])
        if key in persist:
            exc, rea = persist[key]
            row["Exclude"] = exc
            row["Reason"] = rea

    write_xlsx(all_rows, output_path)
    log.info("Processed %d volumes across %d arrays", len(all_rows), len(array_folders))


if __name__ == "__main__":
    main()
