"""Tests for safemode_parser.py against the ashbpasan001 sample data."""

import tempfile
from pathlib import Path

import pytest
from openpyxl import load_workbook

import sys
sys.path.insert(0, str(Path(__file__).parent.parent))

from safemode_parser import parse_array, write_xlsx

SAMPLE_DIR = Path(__file__).parent.parent / "ashbpasan001"

PGROUP_AUTO_MEMBERS = {
    "ashbdr-db-241-A01-n",
    "ashbdr-db-242-A01-n",
    "ashbdr-db-243-A01-n",
    "ashbdr-db-244-A01-n",
    "ashbdr-gc-241-A01-n",
    "ashbdr-gc-242-A01-n",
    "ashbdr-gc-243-A01-n",
    "ashbdr-gc-244-A01-n",
}


@pytest.fixture(scope="module")
def rows():
    return parse_array(SAMPLE_DIR)


def test_volume_count(rows):
    assert len(rows) == 87


def test_safemode_pgroup_member_count(rows):
    safemode = [r for r in rows if r["In SafeMode PGroup"] == "Yes"]
    assert len(safemode) == 8


def test_no_volumes_meet_bp_threshold(rows):
    assert all(r["Max Snap Window"] < 14 for r in rows)


def test_pgroup_auto_snap_window(rows):
    by_name = {r["Volume Name"]: r for r in rows}
    for vol in PGROUP_AUTO_MEMBERS:
        assert by_name[vol]["Max Snap Window"] == 8, f"{vol} expected Max Snap Window=8"


def test_pgroup_auto_safemode_yes(rows):
    by_name = {r["Volume Name"]: r for r in rows}
    for vol in PGROUP_AUTO_MEMBERS:
        assert by_name[vol]["In SafeMode PGroup"] == "Yes", f"{vol} should be In SafeMode PGroup"


def test_risk_level_formula_in_every_row(rows):
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        tmp = Path(f.name)
    try:
        write_xlsx(rows, tmp)
        wb = load_workbook(tmp)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        risk_idx = headers.index("Risk Level")
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            assert row[risk_idx].value is not None, "Risk Level formula missing in a row"
    finally:
        tmp.unlink(missing_ok=True)


def test_missing_out_files_no_crash(tmp_path):
    array_dir = tmp_path / "test_array"
    array_dir.mkdir()
    (array_dir / "purevol-list.out").write_text(
        "Name  Size  Virtual  Unique  Snapshots  Total  Method\n"
        "vol1  10G   1G       1G      0          1G     effective\n"
    )
    result = parse_array(array_dir)
    assert len(result) == 1
    assert result[0]["Volume Name"] == "vol1"
    assert result[0]["Max Snap Window"] == 0
    assert result[0]["In SafeMode PGroup"] == "No"
